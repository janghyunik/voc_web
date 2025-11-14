# convert_birthdays.py
import os, json
from datetime import datetime
from openpyxl import load_workbook

BASE_DIR = os.path.abspath(os.path.dirname(__file__))
XLSX_PATH = os.path.join(BASE_DIR, "birthdays.xlsx")
DATA_DIR = os.path.join(BASE_DIR, "data")
JSON_PATH = os.path.join(DATA_DIR, "birthdays.json")

def parse_month_day(v):
    """excel의 날짜( datetime 또는 문자열 )에서 MM-DD로 변환"""
    if v is None:
        return None
    # datetime 타입인 경우
    if isinstance(v, datetime):
        return f"{v.month:02d}-{v.day:02d}"
    # 문자열인 경우 (예: '2025-11-18 00:00:00' 또는 '2025-11-18')
    s = str(v).strip()
    if not s:
        return None
    # 공백 기준으로 앞부분(날짜)만 취함
    token = s.split()[0]
    # YYYY-MM-DD 또는 MM-DD 지원
    parts = token.split("-")
    try:
        if len(parts) == 3:  # YYYY-MM-DD
            return f"{int(parts[1]):02d}-{int(parts[2]):02d}"
        if len(parts) == 2:  # MM-DD
            return f"{int(parts[0]):02d}-{int(parts[1]):02d}"
    except Exception:
        pass
    # 기타 포맷은 필요 시 추가
    return None

def main():
    if not os.path.isfile(XLSX_PATH):
        raise FileNotFoundError(f"엑셀 파일을 찾을 수 없습니다: {XLSX_PATH}")

    os.makedirs(DATA_DIR, exist_ok=True)

    wb = load_workbook(XLSX_PATH, data_only=True)
    if "Sheet1" not in wb.sheetnames:
        raise ValueError("엑셀에 'Sheet1' 시트를 찾을 수 없습니다.")
    ws = wb["Sheet1"]

    results = []
    # 1행은 헤더라고 가정 → 2행부터 읽기
    for r in range(2, ws.max_row + 1):
        name = ws.cell(row=r, column=1).value  # A열
        bday = ws.cell(row=r, column=2).value  # B열
        if not name:
            continue
        md = parse_month_day(bday)
        if not md:
            continue
        results.append({"name": str(name).strip(), "birthday": md})

    with open(JSON_PATH, "w", encoding="utf-8") as f:
        json.dump(results, f, ensure_ascii=False, indent=2)

    print(f"[OK] {len(results)}건 저장 → {JSON_PATH}")

if __name__ == "__main__":
    main()
